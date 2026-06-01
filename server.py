"""
IDEP — Backend Python / FastAPI
Recibe respuestas del formulario web y las consolida en un Excel centralizado.
"""

from fastapi import FastAPI, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, FileResponse
from fastapi.staticfiles import StaticFiles
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json, os, re
from datetime import datetime
from pathlib import Path

# ─── CONFIG ──────────────────────────────────────────────────────────────────
BASE_DIR   = Path(__file__).parent
DATA_DIR   = BASE_DIR / "data"
EXCEL_FILE = DATA_DIR / "IDEP_Respuestas_Consolidadas.xlsx"
DATA_DIR.mkdir(exist_ok=True)

app = FastAPI(title="IDEP API", version="1.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Serve the frontend (index.html) as static files
app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static"), html=True), name="static")

# ─── COLUMN DEFINITIONS ──────────────────────────────────────────────────────
COLUMNS = [
    ("N°",                                    "meta"),
    ("Fecha y hora",                          "meta_ts"),
    ("Nombre",                                "nombre"),
    ("Rol / Cargo",                           "cargo"),
    ("Email",                                 "email"),
    ("Teléfono",                              "telefono"),
    ("Organización",                          "organizacion"),
    ("Municipio",                             "ciudad"),
    ("Subregión de Caldas",                   "region"),
    ("Tipo de actor (Hélice)",                "actorLabel"),
    ("Tipología específica",                  "typology"),
    ("Ecosistema 1",                          "eco_0"),
    ("Ecosistema 2",                          "eco_1"),
    ("Ecosistema 3",                          "eco_2"),
    # Narrativa
    ("[I] Descripción del ecosistema",        "narrativa_descripcion"),
    ("[I] Nudos críticos",                    "narrativa_nudos"),
    ("[I] Fuerzas positivas",                 "narrativa_fuerzas"),
    ("[I] Articulación de actores",           "narrativa_articulacion"),
    ("[I] Necesidades del ecosistema",        "narrativa_necesidades"),
    ("[I] Oportunidades no aprovechadas",     "narrativa_oportunidades"),
    # Cadenas
    ("[II] Ruptura de la cadena",             "cadena_ruptura"),
    ("[II] Potencial subutilizado",           "cadena_potencial"),
    ("[II] Desconexiones entre actores",      "cadena_desconexion"),
    ("[II] Iniciativa transformadora",        "cadena_iniciativa"),
    # Convergente
    ("[Conv] Articulación actores",           "c1"),
    ("[Conv] Acceso a mercados",              "c2"),
    ("[Conv] Arraigo territorial",            "c3"),
    ("[Conv] Innovación y tecnología",        "c4"),
    ("[Conv] Distribución del valor",         "c5"),
    ("[Conv] Confianza entre actores",        "c6"),
    ("[Conv] Resiliencia y sostenibilidad",   "c7"),
    ("[Conv] Circulación del conocimiento",   "c8"),
    ("[Conv] Diversidad productiva",          "c9"),
    ("[Conv] Coordinación ecosistémica",      "c10"),
    ("[Conv] PUNTAJE TOTAL",                  "conv_total"),
    ("[Conv] NIVEL DE MADUREZ",               "conv_level"),
    # Cierre
    ("[III] Visión 2030",                     "futuro_vision"),
    ("[III] Comentarios adicionales",         "futuro_adicional"),
]

# ─── EXCEL UTILITIES ─────────────────────────────────────────────────────────
C_DARK   = "0C0E18"
C_ACCENT = "152D4A"
C_GOLD   = "B8893A"
C_TEAL   = "1A6E62"
C_MIST   = "F4F1E8"
C_CREAM  = "ECE8DC"
C_WHITE  = "FFFFFF"
C_BORDER = "CCC8BC"

thin   = Side(style="thin",   color=C_BORDER)
medium = Side(style="medium",  color=C_ACCENT)
BORDER_ALL  = Border(left=thin,   right=thin,   top=thin,   bottom=thin)
BORDER_HEAD = Border(left=medium, right=medium, top=medium, bottom=medium)

def _fill(color): return PatternFill("solid", fgColor=color)
def _font(bold=False, size=9, color="111111", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)
def _align(h="left", wrap=True):
    return Alignment(horizontal=h, vertical="top", wrap_text=wrap)

def init_workbook():
    """Create or load the workbook, ensure 'Consolidado' sheet exists."""
    if EXCEL_FILE.exists():
        wb = openpyxl.load_workbook(EXCEL_FILE)
    else:
        wb = openpyxl.Workbook()
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if "Consolidado" not in wb.sheetnames:
        ws = wb.create_sheet("Consolidado", 0)
        _write_header_row(ws)

    return wb

def _write_header_row(ws):
    """Write styled header row to Consolidado sheet."""
    # Row 1: title
    ws.merge_cells(f"A1:{get_column_letter(len(COLUMNS))}1")
    t = ws.cell(row=1, column=1,
                value="IDEP — INSTRUMENTO DIAGNÓSTICO DE ECOSISTEMAS PRODUCTIVOS · Respuestas Consolidadas")
    t.font      = _font(bold=True, size=13, color=C_WHITE)
    t.fill      = _fill(C_DARK)
    t.alignment = _align("center")
    ws.row_dimensions[1].height = 28

    # Row 2: column headers
    for col, (label, _) in enumerate(COLUMNS, 1):
        c = ws.cell(row=2, column=col, value=label)
        c.font      = _font(bold=True, size=8, color=C_WHITE)
        c.fill      = _fill(C_ACCENT)
        c.border    = BORDER_HEAD
        c.alignment = _align("center")
        ws.column_dimensions[get_column_letter(col)].width = max(len(label) + 2, 18)
    ws.row_dimensions[2].height = 36
    ws.freeze_panes = "C3"

def _next_row(ws):
    """Find next empty data row (after header rows 1 & 2)."""
    return max(ws.max_row + 1, 3)

def _maturity_level(score) -> str:
    """Convert numeric score to maturity label."""
    try: score = int(score)
    except: return "—"
    if score <= 30: return "Ecosistema Fragmentado (10–30)"
    if score <= 50: return "Ecosistema Emergente (31–50)"
    if score <= 65: return "Ecosistema Articulado (51–65)"
    if score <= 75: return "Ecosistema Consolidado (66–75)"
    return "Ecosistema Inteligente / Adaptativo (76–80)"

def append_response(data: dict) -> int:
    """Append one response to Consolidado and create individual sheet. Returns row number."""
    wb  = init_workbook()
    ws  = wb["Consolidado"]
    row = _next_row(ws)
    seq = row - 2   # response number

    # Build flat value map
    ecosystems = data.get("ecosystems", [])
    conv       = data.get("convergentScores", {})

    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    values = {
        "meta":                      seq,
        "meta_ts":                   ts,
        "nombre":                    data.get("nombre",""),
        "cargo":                     data.get("cargo",""),
        "email":                     data.get("email",""),
        "telefono":                  data.get("telefono",""),
        "organizacion":              data.get("organizacion",""),
        "ciudad":                    data.get("ciudad",""),
        "region":                    data.get("region",""),
        "actorLabel":                data.get("actorLabel",""),
        "typology":                  data.get("typology",""),
        "eco_0":                     ecosystems[0] if len(ecosystems) > 0 else "",
        "eco_1":                     ecosystems[1] if len(ecosystems) > 1 else "",
        "eco_2":                     ecosystems[2] if len(ecosystems) > 2 else "",
        # Narrativa
        "narrativa_descripcion":     data.get("narrativa_descripcion",""),
        "narrativa_nudos":           data.get("narrativa_nudos",""),
        "narrativa_fuerzas":         data.get("narrativa_fuerzas",""),
        "narrativa_articulacion":    data.get("narrativa_articulacion",""),
        "narrativa_necesidades":     data.get("narrativa_necesidades",""),
        "narrativa_oportunidades":   data.get("narrativa_oportunidades",""),
        # Cadenas
        "cadena_ruptura":            data.get("cadena_ruptura",""),
        "cadena_potencial":          data.get("cadena_potencial",""),
        "cadena_desconexion":        data.get("cadena_desconexion",""),
        "cadena_iniciativa":         data.get("cadena_iniciativa",""),
        # Convergente
        "c1":  str(conv.get("c1","")),  "c2":  str(conv.get("c2","")),
        "c3":  str(conv.get("c3","")),  "c4":  str(conv.get("c4","")),
        "c5":  str(conv.get("c5","")),  "c6":  str(conv.get("c6","")),
        "c7":  str(conv.get("c7","")),  "c8":  str(conv.get("c8","")),
        "c9":  str(conv.get("c9","")),  "c10": str(conv.get("c10","")),
        "conv_total":  data.get("convergentTotal",""),
        "conv_level":  _maturity_level(data.get("convergentTotal",0)),
        # Cierre
        "futuro_vision":             data.get("futuro_vision",""),
        "futuro_adicional":          data.get("futuro_adicional",""),
    }

    # Alternate row colors
    row_fill = _fill(C_MIST) if seq % 2 == 0 else _fill(C_WHITE)

    for col, (_, key) in enumerate(COLUMNS, 1):
        val = values.get(key, "")
        if key == "meta":
            val = seq
            c = ws.cell(row=row, column=col, value=val)
            c.fill      = _fill(C_CREAM)
            c.font      = _font(bold=True, size=9, color=C_ACCENT)
            c.alignment = _align("center", wrap=False)
        else:
            # Second meta column = timestamp
            if col == 2:
                val = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            c = ws.cell(row=row, column=col, value=val)
            # Score cells get teal highlight
            if "_score" in key and val:
                c.fill = _fill("E0F2EF")
                c.font = _font(bold=True, size=9, color=C_TEAL)
            else:
                c.fill = row_fill
                c.font = _font(size=9)
        c.border    = BORDER_ALL
        c.alignment = _align(wrap=True)
        # Row height for text-heavy rows
    ws.row_dimensions[row].height = 55

    # ── Individual sheet per respondent ──────────────────────────────────────
    safe_name = re.sub(r'[\\/*?:\[\]]', '_', data.get("nombre", f"Resp_{seq}"))[:28]
    sheet_name = f"{seq:03d}_{safe_name}"
    if sheet_name in wb.sheetnames:
        sheet_name = f"{seq:03d}_{safe_name[:20]}_{datetime.now().strftime('%H%M%S')}"

    wi = wb.create_sheet(sheet_name)
    _write_individual_sheet(wi, data, seq, values, ecosystems)

    wb.save(EXCEL_FILE)
    return seq

def _write_individual_sheet(ws, data, seq, values, ecosystems):
    """Write a detailed formatted sheet for one respondent."""
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 72

    r = 1

    def title_row(text, fg=C_DARK):
        ws.merge_cells(f"A{r}:B{r}")
        c = ws.cell(row=r, column=1, value=text)
        c.font      = _font(bold=True, size=12, color=C_WHITE)
        c.fill      = _fill(fg)
        c.alignment = _align("center", wrap=False)
        ws.row_dimensions[r].height = 24

    def data_row(label, value, score=False):
        lc = ws.cell(row=r, column=1, value=label)
        vc = ws.cell(row=r, column=2, value=value or "—")
        lc.font      = _font(bold=True, size=9, color="333333")
        lc.fill      = _fill(C_CREAM)
        lc.border    = BORDER_ALL
        lc.alignment = _align()
        vc.font      = _font(bold=score, size=9, color=C_TEAL if score and value else "111111")
        vc.fill      = _fill("E0F2EF") if score and value else _fill(C_WHITE)
        vc.border    = BORDER_ALL
        vc.alignment = _align(wrap=True)
        ws.row_dimensions[r].height = 32 if not score else 22

    def comment_row(value):
        ws.merge_cells(f"A{r}:B{r}")
        c = ws.cell(row=r, column=1, value=f"Ampliación: {value}" if value else "—")
        c.font      = _font(italic=True, size=8, color="555555")
        c.fill      = _fill("FAFAF7")
        c.border    = BORDER_ALL
        c.alignment = _align(wrap=True)
        if value:
            words = len(value.split())
            ws.row_dimensions[r].height = max(30, min(120, words * 1.8))
        else:
            ws.row_dimensions[r].height = 16

    def sep():
        ws.merge_cells(f"A{r}:B{r}")
        ws.cell(row=r, column=1).fill = _fill(C_MIST)
        ws.row_dimensions[r].height = 8

    # Title
    title_row(f"IDEP · Respuesta N° {seq:03d} — {data.get('nombre','')}", C_DARK)
    r += 1
    title_row("Instrumento Diagnóstico de Ecosistemas Productivos · Quinta Hélice · 2025", C_ACCENT)
    r += 1; sep(); r += 1

    # Section 1
    title_row("📋  SECCIÓN 1 · IDENTIFICACIÓN", "1E4A7A")
    r += 1
    for label, key in [("Fecha y hora", None), ("Nombre completo","nombre"),
                        ("Cargo / Rol","cargo"), ("Correo electrónico","email"),
                        ("Teléfono","telefono"), ("Organización","organizacion"),
                        ("Ciudad","ciudad"), ("Departamento","departamento")]:
        val = datetime.now().strftime("%Y-%m-%d %H:%M:%S") if key is None else values.get(key,"")
        data_row(label, val)
        r += 1
    sep(); r += 1

    # Section 2
    title_row("🌀  SECCIÓN 2 · TIPO DE ACTOR", "1A6E62")
    r += 1
    data_row("Hélice / Tipo de actor", values.get("actorLabel",""))
    r += 1
    data_row("Tipología específica",   values.get("typology",""))
    r += 1; sep(); r += 1

    # Section 3
    title_row("🗺️  SECCIÓN 3 · ECOSISTEMAS PRODUCTIVOS", "B8893A")
    r += 1
    for i, eco in enumerate(ecosystems, 1):
        data_row(f"Ecosistema {i}", eco)
        r += 1
    sep(); r += 1

    # Section 4
    title_row("📊  SECCIÓN 4 · ESTADO ACTUAL DEL ECOSISTEMA", "1A3A5C")
    r += 1
    for key_s, key_c, label in [
        ("m1_score","m1_comment","Madurez del ecosistema"),
        ("m2_score","m2_comment","Competitividad sistémica"),
        ("m3_score","m3_comment","Articulación inter-actores"),
        ("m4_score","m4_comment","Capacidad de innovación"),
        ("m5_score","m5_comment","Gobernanza del ecosistema"),
        ("m6_score","m6_comment","Sostenibilidad e impacto territorial"),
    ]:
        data_row(label, values.get(key_s,""), score=True)
        r += 1
        comment_row(values.get(key_c,""))
        r += 1
    sep(); r += 1

    # Section 5
    title_row("🔗  SECCIÓN 5 · MAPEO DE ACTORES, CADENA DE VALOR Y GOBERNANZA", "1A6E62")
    r += 1
    for key_s, key_c, label in [
        ("map1_score","map1_comment","Suficiencia del mapeo de actores"),
        ("map2_score","map2_comment","Articulación en la cadena de valor"),
        ("map3_score","map3_comment","Aplicación práctica en proyectos"),
        ("map4_score","map4_comment","Diagnóstico de gobernanza"),
        ("map5_score","map5_comment","Flujos de información y conocimiento"),
        ("map6_score","map6_comment","Capacidad de análisis estratégico"),
    ]:
        data_row(label, values.get(key_s,""), score=True)
        r += 1
        comment_row(values.get(key_c,""))
        r += 1
    sep(); r += 1

    # Section 6
    title_row("🔬  SECCIÓN 6 · DIAGNÓSTICO DE CADENAS PRODUCTIVAS Y NECESIDADES", "6B2A1E")
    r += 1
    for key_d, label in [
        ("d1","Diagnóstico de cadenas productivas"),
        ("d2","Oportunidades del ecosistema"),
        ("d3","Necesidades urgentes"),
        ("d4","Actores clave y roles"),
        ("d5","Propuestas de mejora"),
        ("d6","Comentarios adicionales"),
    ]:
        data_row(label, "")
        r += 1
        comment_row(values.get(key_d,""))
        r += 1

    sep(); r += 1

    # Section 7: Convergent
    title_row("📊  SECCIÓN 7 · CUESTIONARIO CONVERGENTE DE ECOSISTEMAS PRODUCTIVOS", "1A3A5C")
    r += 1
    conv_labels = [
        ("c1","Articulación entre actores"),("c2","Acceso a mercados"),
        ("c3","Arraigo territorial"),("c4","Desarrollo tecnológico e innovación"),
        ("c5","Distribución del valor"),("c6","Confianza y capital social"),
        ("c7","Sostenibilidad y resiliencia"),("c8","Circulación del conocimiento"),
        ("c9","Base productiva"),("c10","Coordinación ecosistémica"),
    ]
    for key_c, label in conv_labels:
        data_row(label, values.get(key_c,""), score=True)
        r += 1
    # Total row
    ws.merge_cells(f"A{r}:B{r}")
    c = ws.cell(row=r, column=1,
        value=f"PUNTAJE TOTAL: {values.get('conv_total','')} / 80  →  {values.get('conv_level','')}")
    c.font   = _font(bold=True, size=11, color=C_WHITE)
    c.fill   = _fill(C_ACCENT)
    c.border = BORDER_ALL
    c.alignment = _align("center", wrap=False)
    ws.row_dimensions[r].height = 26
    r += 1

# ─── API ENDPOINTS ────────────────────────────────────────────────────────────

@app.get("/")
async def root():
    """Serve the main HTML page."""
    index = BASE_DIR / "static" / "index.html"
    if index.exists():
        return FileResponse(str(index))
    return JSONResponse({"status": "IDEP API running. Place index.html in /static folder."})

@app.post("/api/submit")
async def submit_response(request: Request):
    """Receive form data and save to Excel."""
    try:
        data = await request.json()
        seq  = append_response(data)
        return JSONResponse({
            "ok": True,
            "message": f"Respuesta N° {seq} guardada correctamente.",
            "seq": seq,
            "file": "IDEP_Respuestas_Consolidadas.xlsx"
        })
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

@app.get("/api/download")
async def download_excel():
    """Download the consolidated Excel file."""
    if not EXCEL_FILE.exists():
        return JSONResponse({"error": "No hay respuestas aún."}, status_code=404)
    return FileResponse(
        str(EXCEL_FILE),
        filename="IDEP_Respuestas_Consolidadas.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.get("/api/count")
async def count_responses():
    """Return number of responses collected."""
    if not EXCEL_FILE.exists():
        return {"count": 0}
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
    ws = wb["Consolidado"]
    count = max(ws.max_row - 2, 0)  # subtract 2 header rows
    wb.close()
    return {"count": count}

@app.get("/api/status")
async def status():
    return {"status": "ok", "server": "IDEP API", "version": "1.0",
            "excel": str(EXCEL_FILE), "exists": EXCEL_FILE.exists()}

# ─── MAIN ─────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import uvicorn
    print("\n" + "═"*60)
    print("  IDEP — Servidor Backend")
    print("  http://localhost:8000")
    print("  Admin Excel: GET /api/download")
    print("═"*60 + "\n")
    uvicorn.run("server:app", host="0.0.0.0", port=8000, reload=True)
