"""
IDEP — Instrumento Diagnóstico de Ecosistemas Productivos
Aplicación Python / Streamlit
Investigación Doctoral en Innovación y Productividad Regional
Modelo de la Quinta Hélice
"""

import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
import io
from datetime import date

# ─── PAGE CONFIG ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="IDEP · Diagnóstico de Ecosistemas Productivos",
    page_icon="🌐",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ─── CUSTOM CSS ──────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=DM+Sans:ital,wght@0,300;0,400;0,500;1,300&display=swap');

:root {
    --ink: #0d0f1a;
    --paper: #f5f2eb;
    --accent: #1a3a5c;
    --gold: #c89b3c;
    --teal: #1e7a6e;
    --red: #b03a2e;
    --mist: #e8e4da;
    --mid: #8a8578;
    --border: #d0cbbf;
}

html, body, [class*="css"] {
    font-family: 'DM Sans', sans-serif;
}

.main-header {
    background: linear-gradient(135deg, #0d0f1a 0%, #1a3a5c 60%, #1e7a6e 100%);
    color: white;
    padding: 2.5rem 3rem;
    border-radius: 12px;
    margin-bottom: 2rem;
    position: relative;
    overflow: hidden;
}
.main-header::before {
    content: '';
    position: absolute;
    top: -60px; right: -60px;
    width: 220px; height: 220px;
    background: rgba(200,155,60,0.15);
    border-radius: 50%;
}
.main-header::after {
    content: '';
    position: absolute;
    bottom: -40px; left: 40%;
    width: 160px; height: 160px;
    background: rgba(30,122,110,0.2);
    border-radius: 50%;
}
.badge {
    display: inline-block;
    background: rgba(200,155,60,0.25);
    border: 1px solid rgba(200,155,60,0.6);
    color: #f0c96a;
    padding: 0.3rem 0.9rem;
    border-radius: 100px;
    font-size: 0.72rem;
    font-weight: 700;
    letter-spacing: 0.18em;
    text-transform: uppercase;
    margin-bottom: 1rem;
}
.main-header h1 {
    font-family: 'Syne', sans-serif;
    font-size: 2.2rem;
    font-weight: 800;
    line-height: 1.1;
    margin: 0 0 0.5rem;
    letter-spacing: -0.02em;
}
.main-header p {
    opacity: 0.75;
    font-size: 0.9rem;
    max-width: 600px;
    line-height: 1.6;
    font-weight: 300;
}

.step-header {
    background: #fff;
    border: 1.5px solid #d0cbbf;
    border-left: 5px solid #1a3a5c;
    border-radius: 8px;
    padding: 1.25rem 1.5rem;
    margin-bottom: 1.5rem;
}
.step-header h2 {
    font-family: 'Syne', sans-serif;
    font-size: 1.2rem;
    font-weight: 700;
    color: #1a3a5c;
    margin-bottom: 0.3rem;
}
.step-header p {
    font-size: 0.85rem;
    color: #8a8578;
    line-height: 1.6;
    margin: 0;
    font-style: italic;
}

.ref-box {
    background: rgba(30,122,110,0.06);
    border-left: 3px solid #1e7a6e;
    padding: 0.6rem 1rem;
    border-radius: 0 6px 6px 0;
    font-size: 0.75rem;
    color: #145f56;
    line-height: 1.7;
    margin-bottom: 1.25rem;
}

.actor-card {
    background: white;
    border: 2px solid #d0cbbf;
    border-radius: 10px;
    padding: 1.2rem 0.8rem;
    text-align: center;
    cursor: pointer;
    transition: all 0.2s;
    margin-bottom: 0.5rem;
}
.actor-card:hover { border-color: #1a3a5c; }
.actor-card.selected {
    border-color: #1a3a5c;
    background: #1a3a5c;
    color: white;
}

.scale-card {
    background: white;
    border: 1.5px solid #d0cbbf;
    border-radius: 10px;
    padding: 1.25rem 1.5rem;
    margin-bottom: 1.25rem;
}
.q-tag {
    display: inline-block;
    background: rgba(200,155,60,0.15);
    color: #9a7020;
    padding: 0.2rem 0.6rem;
    border-radius: 4px;
    font-size: 0.7rem;
    font-weight: 700;
    letter-spacing: 0.1em;
    text-transform: uppercase;
    margin-bottom: 0.5rem;
}

.progress-container {
    background: #e8e4da;
    border-radius: 100px;
    height: 8px;
    margin-bottom: 0.5rem;
    overflow: hidden;
}
.progress-bar {
    height: 100%;
    border-radius: 100px;
    background: linear-gradient(90deg, #1a3a5c, #1e7a6e);
    transition: width 0.5s ease;
}

.nav-col button {
    width: 100%;
}

.success-box {
    background: linear-gradient(135deg, #1e7a6e, #145f56);
    color: white;
    border-radius: 12px;
    padding: 2.5rem;
    text-align: center;
    margin: 1rem 0;
}
.success-box h2 {
    font-family: 'Syne', sans-serif;
    font-size: 1.8rem;
    margin-bottom: 0.75rem;
}

.word-counter {
    font-size: 0.72rem;
    color: #8a8578;
    text-align: right;
    margin-top: 0.2rem;
}
.word-counter.over { color: #b03a2e; font-weight: 600; }

.divider-line {
    border: none;
    border-top: 1px solid #d0cbbf;
    margin: 1.5rem 0;
}

/* Streamlit overrides */
.stButton > button {
    font-family: 'Syne', sans-serif;
    font-weight: 700;
    letter-spacing: 0.05em;
    border-radius: 7px;
    padding: 0.55rem 1.5rem;
    transition: all 0.2s;
}
.stSelectbox > div > div {
    border-radius: 7px;
    border-color: #d0cbbf;
}
.stTextInput > div > div > input {
    border-radius: 7px;
    border-color: #d0cbbf;
}
.stTextArea > div > div > textarea {
    border-radius: 7px;
    border-color: #d0cbbf;
    font-family: 'DM Sans', sans-serif;
}
section[data-testid="stSidebar"] { display: none; }
</style>
""", unsafe_allow_html=True)

# ─── DATA ────────────────────────────────────────────────────────────────────

QUINTUPLE_HELIX_ACTORS = {
    "🎓 Academia / Universidad": "academia",
    "🏭 Empresa / Industria": "empresa",
    "🏛️ Gobierno / Estado": "gobierno",
    "🌿 Sociedad Civil / Comunidad": "sociedad",
    "📡 Medios / Cultura / TIC": "media",
}

ACTOR_TYPOLOGIES = {
    "academia": [
        "— Seleccione —",
        "Universidad de investigación intensiva (R1)",
        "Universidad docente con grupos de I+D",
        "Instituto tecnológico / Politécnico",
        "Centro de investigación aplicada",
        "Incubadora o aceleradora universitaria",
        "Laboratorio de innovación abierta (fab lab / living lab)",
        "Red académica interinstitucional",
    ],
    "empresa": [
        "— Seleccione —",
        "Gran empresa / corporación multinacional",
        "Mediana empresa manufacturera o de servicios",
        "PYME innovadora de base tecnológica",
        "Startup o empresa emergente (scale-up)",
        "Empresa social / cooperativa",
        "Clúster o asociación empresarial sectorial",
        "Empresa tractora de cadena de valor",
    ],
    "gobierno": [
        "— Seleccione —",
        "Ministerio o secretaría de ciencia, tecnología e innovación",
        "Agencia de desarrollo regional o territorial",
        "Gobierno local / alcaldía / municipio",
        "Entidad regulatoria o de política pública",
        "Fondo de financiamiento público (CONACYT, Minciencias, etc.)",
        "Parque tecnológico o zona económica especial",
        "Corporación de desarrollo económico",
    ],
    "sociedad": [
        "— Seleccione —",
        "Organización de la sociedad civil (ONG / fundación)",
        "Comunidad indígena o campesina organizada",
        "Asociación de consumidores o usuarios",
        "Red de innovación social y economía solidaria",
        "Sindicato o gremio laboral",
        "Colectivo ciudadano de innovación participativa",
        "Organismo multilateral / cooperación internacional",
    ],
    "media": [
        "— Seleccione —",
        "Plataforma digital o ecosistema de datos abiertos",
        "Medio de comunicación especializado en innovación",
        "Red social / comunidad en línea de práctica",
        "Centro de inteligencia artificial o cómputo avanzado",
        "Empresa de telecomunicaciones o proveedor TIC",
        "Organización de gestión del conocimiento y vigilancia tecnológica",
        "Consorcio de transferencia tecnológica y propiedad intelectual",
    ],
}

PRODUCTIVE_ECOSYSTEMS = [
    "Ecosistema agroindustrial y bioeconomía",
    "Ecosistema de manufactura avanzada e industria 4.0",
    "Ecosistema de tecnologías de información y software",
    "Ecosistema de energías renovables y transición energética",
    "Ecosistema de salud, farmacéutica y biotecnología",
    "Ecosistema de turismo sostenible y economía creativa",
    "Ecosistema de logística, transporte y cadena de suministro",
    "Ecosistema minero-energético y recursos naturales",
    "Ecosistema de construcción, infraestructura y ciudades inteligentes",
    "Ecosistema de economía circular y gestión ambiental",
    "Ecosistema financiero y fintech",
    "Ecosistema educativo y edtech",
    "Ecosistema de defensa, seguridad y tecnologías duales",
    "Ecosistema de textil, moda y diseño industrial",
    "Ecosistema de alimentos funcionales y nutrición",
]

MATURITY_QUESTIONS = [
    {
        "id": "m1", "label": "Madurez del ecosistema",
        "question": "¿Cómo evalúa el nivel de madurez general del ecosistema productivo regional en términos de capacidades instaladas, infraestructura y trayectoria de desarrollo?",
        "scale": ["1 – Incipiente", "2 – En formación", "3 – En desarrollo", "4 – Consolidado", "5 – Avanzado/Líder"],
    },
    {
        "id": "m2", "label": "Competitividad sistémica",
        "question": "¿En qué medida considera que el ecosistema regional es competitivo frente a ecosistemas equivalentes a nivel nacional e internacional?",
        "scale": ["1 – Muy baja", "2 – Baja", "3 – Media", "4 – Alta", "5 – Muy alta"],
    },
    {
        "id": "m3", "label": "Articulación inter-actores",
        "question": "¿Qué tan efectiva es la articulación entre los diferentes actores del ecosistema (academia, empresa, gobierno, sociedad civil y medios/TIC)?",
        "scale": ["1 – Nula", "2 – Escasa", "3 – Parcial", "4 – Frecuente", "5 – Sistémica"],
    },
    {
        "id": "m4", "label": "Capacidad de innovación",
        "question": "¿Cómo valora la capacidad del ecosistema para generar, adoptar y escalar innovaciones de producto, proceso, organización o mercado?",
        "scale": ["1 – Muy limitada", "2 – Limitada", "3 – Moderada", "4 – Sólida", "5 – Excelente"],
    },
    {
        "id": "m5", "label": "Gobernanza del ecosistema",
        "question": "¿Existe una gobernanza clara y legítima que facilite la toma de decisiones colectivas, la gestión de conflictos y la orientación estratégica del ecosistema?",
        "scale": ["1 – Inexistente", "2 – Informal/Débil", "3 – Emergente", "4 – Funcional", "5 – Institucionalizada"],
    },
    {
        "id": "m6", "label": "Sostenibilidad e impacto territorial",
        "question": "¿En qué medida las actividades del ecosistema generan impacto positivo, sostenible y equitativo en el territorio regional?",
        "scale": ["1 – Sin impacto visible", "2 – Impacto puntual", "3 – Impacto moderado", "4 – Impacto significativo", "5 – Impacto transformador"],
    },
]

MAPPING_QUESTIONS = [
    {
        "id": "map1", "label": "Suficiencia del mapeo de actores",
        "question": "¿Considera que existe información suficiente, actualizada y accesible sobre los actores que conforman el ecosistema productivo regional?",
        "scale": ["1 – Totalmente insuficiente", "2 – Insuficiente", "3 – Parcialmente suficiente", "4 – Suficiente", "5 – Muy completa"],
    },
    {
        "id": "map2", "label": "Articulación en la cadena de valor",
        "question": "¿Qué tan articulados están los eslabones de la cadena de valor regional (proveedores, productores, transformadores, distribuidores, consumidores finales)?",
        "scale": ["1 – Desarticulados", "2 – Con vínculos esporádicos", "3 – Con vínculos parciales", "4 – Bien articulados", "5 – Totalmente integrados"],
    },
    {
        "id": "map3", "label": "Aplicación práctica en proyectos",
        "question": "¿En qué medida el conocimiento sobre el ecosistema se traduce en proyectos concretos de mejora de competitividad, productividad o innovación?",
        "scale": ["1 – Sin aplicación", "2 – Aplicación esporádica", "3 – Aplicación parcial", "4 – Aplicación sistemática", "5 – Integración estratégica"],
    },
    {
        "id": "map4", "label": "Diagnóstico de gobernanza",
        "question": "¿Existen mecanismos formales o informales de diagnóstico continuo de la gobernanza del ecosistema que permitan identificar brechas y tomar decisiones informadas?",
        "scale": ["1 – No existen", "2 – Muy limitados", "3 – Parcialmente presentes", "4 – Presentes y funcionales", "5 – Robustos y sistemáticos"],
    },
    {
        "id": "map5", "label": "Flujos de información y conocimiento",
        "question": "¿Cómo evalúa la calidad y fluidez de los flujos de información y conocimiento entre los actores del ecosistema para la toma de decisiones colaborativas?",
        "scale": ["1 – Muy deficientes", "2 – Deficientes", "3 – Regulares", "4 – Buenos", "5 – Excelentes"],
    },
    {
        "id": "map6", "label": "Capacidad de análisis estratégico colectivo",
        "question": "¿Tiene el ecosistema la capacidad de analizar colectivamente su posición competitiva, identificar oportunidades y formular estrategias territoriales compartidas?",
        "scale": ["1 – Nula", "2 – Muy baja", "3 – Emergente", "4 – Desarrollada", "5 – Consolidada"],
    },
]

DIAGNOSIS_QUESTIONS = [
    {
        "id": "d1", "label": "Diagnóstico de cadenas productivas",
        "question": "¿Cuáles son las principales brechas, cuellos de botella y oportunidades que usted identifica en las cadenas productivas regionales?",
    },
    {
        "id": "d2", "label": "Oportunidades del ecosistema",
        "question": "¿Qué oportunidades estratégicas —tecnológicas, de mercado, de cooperación o de política pública— identifica para fortalecer el ecosistema productivo regional?",
    },
    {
        "id": "d3", "label": "Necesidades urgentes del ecosistema",
        "question": "¿Cuáles son las necesidades más urgentes que deben atenderse para mejorar la competitividad y articulación del ecosistema productivo?",
    },
    {
        "id": "d4", "label": "Actores clave y roles",
        "question": "¿Quiénes son los actores que considera más influyentes o estratégicos en el ecosistema y qué roles cumplen actualmente?",
    },
    {
        "id": "d5", "label": "Propuestas de mejora",
        "question": "Desde su posición como actor del ecosistema, ¿qué acciones, proyectos o iniciativas propondría para fortalecer la productividad y la innovación regional?",
    },
    {
        "id": "d6", "label": "Comentarios adicionales",
        "question": "¿Desea agregar algún comentario, observación o información complementaria relevante para el diagnóstico del ecosistema productivo regional?",
    },
]

STEPS = [
    "Identificación",
    "Tipo de Actor",
    "Tipología",
    "Ecosistemas",
    "Estado Actual",
    "Mapeo & Gobernanza",
    "Diagnóstico Final",
]

# ─── SESSION STATE INIT ───────────────────────────────────────────────────────
def init_state():
    defaults = {
        "step": 0,
        "submitted": False,
        # Step 0
        "nombre": "", "cargo": "", "email": "", "telefono": "",
        "organizacion": "", "ciudad": "", "departamento": "",
        "fecha": str(date.today()),
        # Step 1
        "actor_type_label": None,
        "actor_type_id": None,
        # Step 2
        "typology": "— Seleccione —",
        # Step 3
        "ecosystems": [],
        # Step 4
        "maturity_scores": {},
        "maturity_comments": {},
        # Step 5
        "mapping_scores": {},
        "mapping_comments": {},
        # Step 6
        "diagnosis": {},
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

init_state()

# ─── HELPERS ─────────────────────────────────────────────────────────────────
def word_count(text):
    if not text:
        return 0
    return len(text.strip().split())

def word_counter_html(text, max_w=200):
    w = word_count(text)
    cls = "over" if w > max_w else ""
    warn = " ⚠ Límite excedido" if w > max_w else ""
    return f'<div class="word-counter {cls}">{w} / {max_w} palabras{warn}</div>'

def progress_html(current, total):
    pct = int((current / total) * 100)
    return f"""
    <div class="progress-container">
        <div class="progress-bar" style="width:{pct}%"></div>
    </div>
    <p style="font-size:0.75rem;color:#8a8578;margin-bottom:1rem;">
        Paso {current} de {total}: <strong>{STEPS[current-1]}</strong>
    </p>
    """

# ─── EXCEL EXPORT ────────────────────────────────────────────────────────────
def generate_excel():
    wb = openpyxl.Workbook()

    # ── Sheet 1: Respuesta completa ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Respuesta IDEP"

    # Color palette
    C_DARK    = "0D0F1A"
    C_ACCENT  = "1A3A5C"
    C_GOLD    = "C89B3C"
    C_TEAL    = "1E7A6E"
    C_MIST    = "F5F2EB"
    C_LIGHT   = "FFFFFF"
    C_BORDER  = "D0CBBF"

    fill_dark   = PatternFill("solid", fgColor=C_DARK)
    fill_accent = PatternFill("solid", fgColor=C_ACCENT)
    fill_gold   = PatternFill("solid", fgColor=C_GOLD)
    fill_teal   = PatternFill("solid", fgColor=C_TEAL)
    fill_mist   = PatternFill("solid", fgColor="E8E4DA")
    fill_light  = PatternFill("solid", fgColor=C_LIGHT)
    fill_section= PatternFill("solid", fgColor="EAF4F2")

    thin = Side(style="thin", color=C_BORDER)
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    font_title  = Font(name="Calibri", bold=True, size=16, color=C_LIGHT)
    font_sub    = Font(name="Calibri", size=10, color="D0C090", italic=True)
    font_header = Font(name="Calibri", bold=True, size=10, color=C_LIGHT)
    font_section= Font(name="Calibri", bold=True, size=11, color=C_ACCENT)
    font_label  = Font(name="Calibri", bold=True, size=9, color="444444")
    font_value  = Font(name="Calibri", size=9, color="111111")
    font_score  = Font(name="Calibri", bold=True, size=10, color=C_TEAL)
    font_ref    = Font(name="Calibri", size=8, color="888888", italic=True)

    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def section_row(ws, row, title, fill):
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row=row, column=1, value=title)
        c.font = font_section
        c.fill = fill
        c.alignment = center
        c.border = border_all
        ws.row_dimensions[row].height = 22

    def data_row(ws, row, label, value, score=False):
        lc = ws.cell(row=row, column=1, value=label)
        vc = ws.cell(row=row, column=2, value=value)
        lc.font = font_label
        vc.font = font_score if score else font_value
        lc.fill = fill_mist
        vc.fill = fill_light
        lc.border = border_all
        vc.border = border_all
        lc.alignment = Alignment(vertical="top", wrap_text=True)
        vc.alignment = Alignment(vertical="top", wrap_text=True)
        ws.merge_cells(f"B{row}:D{row}")
        ws.row_dimensions[row].height = 30

    def comment_row(ws, row, comment):
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row=row, column=1, value=f"💬 {comment}" if comment else "—")
        c.font = font_value
        c.fill = PatternFill("solid", fgColor="FAFAF7")
        c.border = border_all
        c.alignment = Alignment(vertical="top", wrap_text=True)
        if comment:
            ws.row_dimensions[row].height = max(40, min(120, word_count(comment) * 2))
        else:
            ws.row_dimensions[row].height = 18

    # Set column widths
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    # ── TITLE BLOCK ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:D1")
    t = ws.cell(row=1, column=1, value="IDEP — Instrumento Diagnóstico de Ecosistemas Productivos")
    t.font = Font(name="Calibri", bold=True, size=18, color=C_LIGHT)
    t.fill = fill_dark
    t.alignment = center
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:D2")
    s = ws.cell(row=2, column=1, value="Investigación Doctoral en Innovación y Productividad · Modelo de la Quinta Hélice")
    s.font = font_sub
    s.fill = fill_accent
    s.alignment = center
    ws.row_dimensions[2].height = 18

    ws.merge_cells("A3:D3")
    d = ws.cell(row=3, column=1, value=f"Fecha de diligenciamiento: {st.session_state.fecha}")
    d.font = Font(name="Calibri", size=9, color="AAAAAA")
    d.fill = fill_dark
    d.alignment = center
    ws.row_dimensions[3].height = 16

    row = 5

    # ── SECTION 1: Identificación ────────────────────────────────────────────
    section_row(ws, row, "📋  SECCIÓN 1 · IDENTIFICACIÓN DEL ACTOR", PatternFill("solid", fgColor="E8EFF6"))
    row += 1
    fields_id = [
        ("Nombre completo", st.session_state.nombre),
        ("Cargo / Rol", st.session_state.cargo),
        ("Correo electrónico", st.session_state.email),
        ("Teléfono / WhatsApp", st.session_state.telefono),
        ("Organización", st.session_state.organizacion),
        ("Ciudad", st.session_state.ciudad),
        ("Departamento / Región", st.session_state.departamento),
    ]
    for label, value in fields_id:
        data_row(ws, row, label, value)
        row += 1

    row += 1

    # ── SECTION 2: Tipo de Actor ─────────────────────────────────────────────
    section_row(ws, row, "🌀  SECCIÓN 2 · TIPO DE ACTOR · Modelo de la Quinta Hélice", PatternFill("solid", fgColor="EAF4F2"))
    row += 1
    data_row(ws, row, "Hélice / Tipo de actor", st.session_state.actor_type_label or "—")
    row += 1
    data_row(ws, row, "Tipología específica", st.session_state.typology if st.session_state.typology != "— Seleccione —" else "—")
    row += 1

    row += 1

    # ── SECTION 3: Ecosistemas ───────────────────────────────────────────────
    section_row(ws, row, "🗺️  SECCIÓN 3 · ECOSISTEMAS PRODUCTIVOS REGIONALES", PatternFill("solid", fgColor="F5F0E8"))
    row += 1
    for i, eco in enumerate(st.session_state.ecosystems, 1):
        data_row(ws, row, f"Ecosistema {i}", eco)
        row += 1

    row += 1

    # ── SECTION 4: Madurez & Competitividad ─────────────────────────────────
    section_row(ws, row, "📊  SECCIÓN 4 · ESTADO ACTUAL DEL ECOSISTEMA (Madurez, Competitividad, Articulación)", PatternFill("solid", fgColor="E8EEF5"))
    row += 1
    for q in MATURITY_QUESTIONS:
        data_row(ws, row, q["label"], st.session_state.maturity_scores.get(q["id"], "—"), score=True)
        row += 1
        comment_row(ws, row, st.session_state.maturity_comments.get(q["id"], ""))
        row += 1

    row += 1

    # ── SECTION 5: Mapeo & Gobernanza ────────────────────────────────────────
    section_row(ws, row, "🔗  SECCIÓN 5 · MAPEO DE ACTORES, CADENA DE VALOR Y GOBERNANZA", PatternFill("solid", fgColor="EAF4F2"))
    row += 1
    for q in MAPPING_QUESTIONS:
        data_row(ws, row, q["label"], st.session_state.mapping_scores.get(q["id"], "—"), score=True)
        row += 1
        comment_row(ws, row, st.session_state.mapping_comments.get(q["id"], ""))
        row += 1

    row += 1

    # ── SECTION 6: Diagnóstico ────────────────────────────────────────────────
    section_row(ws, row, "🔬  SECCIÓN 6 · DIAGNÓSTICO DE CADENAS PRODUCTIVAS Y NECESIDADES DEL ECOSISTEMA", PatternFill("solid", fgColor="F5F2EB"))
    row += 1
    for q in DIAGNOSIS_QUESTIONS:
        data_row(ws, row, q["label"], "")
        row += 1
        comment_row(ws, row, st.session_state.diagnosis.get(q["id"], ""))
        row += 1

    row += 2

    # ── REFERENCES ───────────────────────────────────────────────────────────
    section_row(ws, row, "📚  REFERENCIAS BIBLIOGRÁFICAS (Q1/Q2)", PatternFill("solid", fgColor="EEEEEE"))
    row += 1
    refs = [
        "Carayannis, E. G., & Campbell, D. F. J. (2010). Triple Helix, Quadruple Helix and Quintuple Helix. Int. J. Social Ecology and Sustainable Development, 1(1), 41–69.",
        "Jacobides, M. G., Cennamo, C., & Gawer, A. (2018). Towards a theory of ecosystems. Strategic Management Journal, 39(8), 2255–2276.",
        "Granstrand, O., & Holgersson, M. (2020). Innovation ecosystems. Technovation, 90–91, 102098.",
        "Autio, E., Nambisan, S., Thomas, L. D. W., & Wright, M. (2018). Digital affordances and entrepreneurial ecosystems. Strategic Entrepreneurship Journal, 12(1), 72–95.",
        "Gereffi, G., & Fernandez-Stark, K. (2016). Global Value Chain Analysis: A Primer. CGGC, Duke University.",
        "Stam, E. (2015). Entrepreneurial ecosystems and regional policy. European Planning Studies, 23(6), 1759–1762.",
        "Lundvall, B.-Å. (2010). National Systems of Innovation. Anthem Press.",
    ]
    for ref in refs:
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row=row, column=1, value=ref)
        c.font = font_ref
        c.fill = PatternFill("solid", fgColor="F8F8F5")
        c.border = border_all
        c.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[row].height = 28
        row += 1

    # ── Sheet 2: Datos planos para análisis ─────────────────────────────────
    ws2 = wb.create_sheet("Datos para Análisis")
    headers = ["Fecha", "Nombre", "Cargo", "Email", "Teléfono", "Organización", "Ciudad", "Departamento",
               "Tipo Actor", "Tipología"]
    for i, eco in enumerate(["Ecosistema 1", "Ecosistema 2", "Ecosistema 3"]):
        headers.append(eco)
    for q in MATURITY_QUESTIONS:
        headers.append(f"[Estado] {q['label']} – Puntaje")
        headers.append(f"[Estado] {q['label']} – Ampliación")
    for q in MAPPING_QUESTIONS:
        headers.append(f"[Mapeo] {q['label']} – Puntaje")
        headers.append(f"[Mapeo] {q['label']} – Ampliación")
    for q in DIAGNOSIS_QUESTIONS:
        headers.append(f"[Diagnóstico] {q['label']}")

    for col, h in enumerate(headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = Font(name="Calibri", bold=True, size=9, color=C_LIGHT)
        c.fill = fill_accent
        c.border = border_all
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws2.column_dimensions[get_column_letter(col)].width = max(18, len(h) * 0.8)
    ws2.row_dimensions[1].height = 40

    ecos_padded = st.session_state.ecosystems + ["", "", ""]

    row_data = [
        st.session_state.fecha,
        st.session_state.nombre,
        st.session_state.cargo,
        st.session_state.email,
        st.session_state.telefono,
        st.session_state.organizacion,
        st.session_state.ciudad,
        st.session_state.departamento,
        st.session_state.actor_type_label or "",
        st.session_state.typology if st.session_state.typology != "— Seleccione —" else "",
        ecos_padded[0], ecos_padded[1], ecos_padded[2],
    ]
    for q in MATURITY_QUESTIONS:
        row_data.append(st.session_state.maturity_scores.get(q["id"], ""))
        row_data.append(st.session_state.maturity_comments.get(q["id"], ""))
    for q in MAPPING_QUESTIONS:
        row_data.append(st.session_state.mapping_scores.get(q["id"], ""))
        row_data.append(st.session_state.mapping_comments.get(q["id"], ""))
    for q in DIAGNOSIS_QUESTIONS:
        row_data.append(st.session_state.diagnosis.get(q["id"], ""))

    for col, val in enumerate(row_data, 1):
        c = ws2.cell(row=2, column=col, value=val)
        c.font = font_value
        c.border = border_all
        c.alignment = Alignment(vertical="top", wrap_text=True)
        if col in [1]:
            c.fill = fill_mist
        else:
            c.fill = fill_light
    ws2.row_dimensions[2].height = 80

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ─── HEADER ──────────────────────────────────────────────────────────────────
st.markdown("""
<div class="main-header">
    <div class="badge">🌐 Quinta Hélice · Diagnóstico Regional · 2025</div>
    <h1>Instrumento Diagnóstico de<br>Ecosistemas Productivos</h1>
    <p>IDEP — Investigación Doctoral en Innovación y Productividad Regional.<br>
    Herramienta de levantamiento de capacidades del ecosistema productivo basada en el modelo de la Quinta Hélice.</p>
</div>
""", unsafe_allow_html=True)

# ─── PROGRESS ────────────────────────────────────────────────────────────────
if not st.session_state.submitted:
    st.markdown(progress_html(st.session_state.step + 1, len(STEPS)), unsafe_allow_html=True)

# ─── SUBMITTED SCREEN ────────────────────────────────────────────────────────
if st.session_state.submitted:
    st.markdown(f"""
    <div class="success-box">
        <div style="font-size:3.5rem;margin-bottom:0.75rem;">✅</div>
        <h2>¡Formulario completado!</h2>
        <p style="opacity:0.85;font-size:0.95rem;max-width:500px;margin:0 auto;">
            Gracias, <strong>{st.session_state.nombre}</strong>. Su información ha sido registrada y 
            contribuirá al diagnóstico del ecosistema productivo regional.<br>
            Descargue sus respuestas en el archivo Excel a continuación.
        </p>
    </div>
    """, unsafe_allow_html=True)

    excel_bytes = generate_excel()
    filename = f"IDEP_{st.session_state.nombre.replace(' ', '_')}_{st.session_state.fecha}.xlsx"
    st.download_button(
        label="📥  Descargar datos en Excel (.xlsx)",
        data=excel_bytes,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    st.markdown("<br>", unsafe_allow_html=True)
    if st.button("↩  Iniciar nueva respuesta", use_container_width=False):
        for key in list(st.session_state.keys()):
            del st.session_state[key]
        st.rerun()
    st.stop()

# ─── STEP 0: IDENTIFICACIÓN ──────────────────────────────────────────────────
if st.session_state.step == 0:
    st.markdown("""
    <div class="step-header">
        <h2>📋 Paso 1 · Datos de Identificación</h2>
        <p>Complete sus datos personales y de la organización. Esta información es confidencial y de uso exclusivo para el diagnóstico del ecosistema productivo regional.</p>
    </div>
    """, unsafe_allow_html=True)

    c1, c2 = st.columns(2)
    with c1:
        st.session_state.nombre = st.text_input("Nombre completo *", value=st.session_state.nombre, placeholder="Ej: María García López")
        st.session_state.email = st.text_input("Correo electrónico *", value=st.session_state.email, placeholder="correo@organizacion.com")
        st.session_state.organizacion = st.text_input("Nombre de la organización *", value=st.session_state.organizacion, placeholder="Ej: Universidad Industrial de Santander")
        st.session_state.ciudad = st.text_input("Ciudad", value=st.session_state.ciudad, placeholder="Ej: Bucaramanga")
    with c2:
        st.session_state.cargo = st.text_input("Cargo / Rol en la organización", value=st.session_state.cargo, placeholder="Ej: Directora de Innovación")
        st.session_state.telefono = st.text_input("Teléfono / WhatsApp", value=st.session_state.telefono, placeholder="+57 300 000 0000")
        st.session_state.departamento = st.text_input("Departamento / Región", value=st.session_state.departamento, placeholder="Ej: Santander")
        st.text_input("Fecha de diligenciamiento", value=st.session_state.fecha, disabled=True)

    can_advance = bool(st.session_state.nombre and st.session_state.email and st.session_state.organizacion)
    col_nav = st.columns([3, 1])
    with col_nav[1]:
        if st.button("Siguiente →", disabled=not can_advance, use_container_width=True, type="primary"):
            st.session_state.step = 1
            st.rerun()

# ─── STEP 1: TIPO DE ACTOR ────────────────────────────────────────────────────
elif st.session_state.step == 1:
    st.markdown("""
    <div class="step-header">
        <h2>🌀 Paso 2 · Tipo de Actor · Modelo de la Quinta Hélice</h2>
        <p>Seleccione el tipo de actor del ecosistema con el que se identifica su organización.</p>
    </div>
    <div class="ref-box">
        <strong>Referencias Q1/Q2:</strong> Carayannis, E. G., Barth, T. D., & Campbell, D. F. (2012). The Quintuple Helix innovation model. <em>Journal of Innovation and Entrepreneurship, 1</em>(1), 2. · Leydesdorff, L. (2012). The Triple Helix, Quadruple Helix, and N-Tuple of Helices. <em>Journal of the Knowledge Economy, 3</em>, 25–35.
    </div>
    """, unsafe_allow_html=True)

    cols = st.columns(len(QUINTUPLE_HELIX_ACTORS))
    for col, (label, actor_id) in zip(cols, QUINTUPLE_HELIX_ACTORS.items()):
        with col:
            selected = st.session_state.actor_type_label == label
            btn_type = "primary" if selected else "secondary"
            if st.button(label, key=f"actor_{actor_id}", use_container_width=True, type=btn_type):
                st.session_state.actor_type_label = label
                st.session_state.actor_type_id = actor_id
                st.session_state.typology = "— Seleccione —"
                st.rerun()

    if st.session_state.actor_type_label:
        st.success(f"✓ Seleccionado: **{st.session_state.actor_type_label}**")

    col_nav1, col_nav2, col_nav3 = st.columns([1, 2, 1])
    with col_nav1:
        if st.button("← Anterior", use_container_width=True):
            st.session_state.step = 0
            st.rerun()
    with col_nav3:
        can = bool(st.session_state.actor_type_id)
        if st.button("Siguiente →", disabled=not can, use_container_width=True, type="primary"):
            st.session_state.step = 2
            st.rerun()

# ─── STEP 2: TIPOLOGÍA ────────────────────────────────────────────────────────
elif st.session_state.step == 2:
    st.markdown(f"""
    <div class="step-header">
        <h2>🔍 Paso 3 · Tipología Específica del Actor</h2>
        <p>Dentro de la categoría <strong>{st.session_state.actor_type_label}</strong>, seleccione la tipología que mejor describe su organización.</p>
    </div>
    <div class="ref-box">
        <strong>Referencias Q1/Q2:</strong> Lundvall, B.-Å. (2010). <em>National Systems of Innovation.</em> Anthem Press. · Autio, E. et al. (2018). Digital affordances, spatial affordances, and the genesis of entrepreneurial ecosystems. <em>Strategic Entrepreneurship Journal, 12</em>(1), 72–95.
    </div>
    """, unsafe_allow_html=True)

    typologies = ACTOR_TYPOLOGIES.get(st.session_state.actor_type_id, [])
    st.session_state.typology = st.selectbox(
        "Seleccione su tipología *",
        options=typologies,
        index=typologies.index(st.session_state.typology) if st.session_state.typology in typologies else 0,
    )

    if st.session_state.typology != "— Seleccione —":
        st.info(f"✓ Tipología seleccionada: **{st.session_state.typology}**")

    col_nav1, _, col_nav3 = st.columns([1, 2, 1])
    with col_nav1:
        if st.button("← Anterior", use_container_width=True):
            st.session_state.step = 1
            st.rerun()
    with col_nav3:
        can = st.session_state.typology != "— Seleccione —"
        if st.button("Siguiente →", disabled=not can, use_container_width=True, type="primary"):
            st.session_state.step = 3
            st.rerun()

# ─── STEP 3: ECOSISTEMAS ─────────────────────────────────────────────────────
elif st.session_state.step == 3:
    st.markdown("""
    <div class="step-header">
        <h2>🗺️ Paso 4 · Ecosistemas Productivos Regionales</h2>
        <p>Identifique en cuáles ecosistemas productivos regionales tiene incidencia su organización (máximo 3 opciones).</p>
    </div>
    <div class="ref-box">
        <strong>Referencias Q1/Q2:</strong> Jacobides, M. G., Cennamo, C., & Gawer, A. (2018). Towards a theory of ecosystems. <em>Strategic Management Journal, 39</em>(8), 2255–2276. · Granstrand, O., & Holgersson, M. (2020). Innovation ecosystems: A conceptual review and a new definition. <em>Technovation, 90–91</em>, 102098.
    </div>
    """, unsafe_allow_html=True)

    selected_count = len(st.session_state.ecosystems)
    st.markdown(f"**Seleccionados: {selected_count}/3** — {'✅ Máximo alcanzado' if selected_count >= 3 else 'Seleccione hasta 3 ecosistemas'}")

    cols = st.columns(3)
    for i, eco in enumerate(PRODUCTIVE_ECOSYSTEMS):
        col = cols[i % 3]
        with col:
            checked = eco in st.session_state.ecosystems
            disabled = (not checked) and (selected_count >= 3)
            label_display = f"{'✅ ' if checked else ''}{eco}"
            if st.checkbox(
                eco,
                value=checked,
                key=f"eco_{i}",
                disabled=disabled,
            ):
                if eco not in st.session_state.ecosystems:
                    st.session_state.ecosystems.append(eco)
                    st.rerun()
            else:
                if eco in st.session_state.ecosystems:
                    st.session_state.ecosystems.remove(eco)
                    st.rerun()

    col_nav1, _, col_nav3 = st.columns([1, 2, 1])
    with col_nav1:
        if st.button("← Anterior", use_container_width=True):
            st.session_state.step = 2
            st.rerun()
    with col_nav3:
        can = len(st.session_state.ecosystems) >= 1
        if st.button("Siguiente →", disabled=not can, use_container_width=True, type="primary"):
            st.session_state.step = 4
            st.rerun()

# ─── STEP 4: MADUREZ & COMPETITIVIDAD ────────────────────────────────────────
elif st.session_state.step == 4:
    st.markdown("""
    <div class="step-header">
        <h2>📊 Paso 5 · Estado Actual del Ecosistema Productivo</h2>
        <p>Evalúe los siguientes aspectos desde su perspectiva como actor. Amplíe con comentarios si lo considera pertinente (máx. 200 palabras).</p>
    </div>
    <div class="ref-box">
        <strong>Referencias Q1/Q2:</strong> Isenberg, D. J. (2011). The entrepreneurship ecosystem strategy as a new paradigm for economy policy. IIEA. · Stam, E., & Spigel, B. (2018). Entrepreneurial Ecosystems. <em>SAGE Handbook for Entrepreneurship and Small Business.</em> · Solvell, O. (2015). <em>Clusters: Balancing Evolutionary and Constructive Forces.</em> Ivory Tower.
    </div>
    """, unsafe_allow_html=True)

    all_answered = True
    for q in MATURITY_QUESTIONS:
        st.markdown(f"""
        <div class="scale-card">
            <div class="q-tag">{q['label']}</div>
        """, unsafe_allow_html=True)
        st.markdown(f"**{q['question']}**")

        current_val = st.session_state.maturity_scores.get(q["id"])
        idx = q["scale"].index(current_val) if current_val in q["scale"] else None
        chosen = st.radio(
            "Puntuación:",
            options=q["scale"],
            index=idx,
            key=f"mat_score_{q['id']}",
            horizontal=True,
        )
        if chosen:
            st.session_state.maturity_scores[q["id"]] = chosen
        else:
            all_answered = False

        comment = st.text_area(
            "Amplíe su respuesta (máx. 200 palabras):",
            value=st.session_state.maturity_comments.get(q["id"], ""),
            key=f"mat_comment_{q['id']}",
            placeholder="Agregue ejemplos, contexto o perspectiva personal...",
            height=90,
        )
        st.session_state.maturity_comments[q["id"]] = comment
        st.markdown(word_counter_html(comment), unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)
        st.markdown("---")

    all_answered = all(
        st.session_state.maturity_scores.get(q["id"]) in q["scale"]
        for q in MATURITY_QUESTIONS
    )

    col_nav1, _, col_nav3 = st.columns([1, 2, 1])
    with col_nav1:
        if st.button("← Anterior", use_container_width=True):
            st.session_state.step = 3
            st.rerun()
    with col_nav3:
        if st.button("Siguiente →", disabled=not all_answered, use_container_width=True, type="primary"):
            st.session_state.step = 5
            st.rerun()

# ─── STEP 5: MAPEO & GOBERNANZA ──────────────────────────────────────────────
elif st.session_state.step == 5:
    st.markdown("""
    <div class="step-header">
        <h2>🔗 Paso 6 · Mapeo de Actores, Cadena de Valor y Gobernanza</h2>
        <p>Evalúe los niveles de suficiencia, articulación y aplicación práctica desde las dimensiones de mapeo, cadena de valor y diagnóstico de gobernanza.</p>
    </div>
    <div class="ref-box">
        <strong>Referencias Q1/Q2:</strong> Kaplinsky, R., & Morris, M. (2001). <em>A handbook for value chain research.</em> IDRC. · Gereffi, G., & Fernandez-Stark, K. (2016). <em>Global Value Chain Analysis: A Primer.</em> CGGC. · Rhodes, R. A. W. (1996). The new governance. <em>Political Studies, 44</em>(4), 652–667.
    </div>
    """, unsafe_allow_html=True)

    for q in MAPPING_QUESTIONS:
        st.markdown(f"""
        <div class="scale-card">
            <div class="q-tag">{q['label']}</div>
        """, unsafe_allow_html=True)
        st.markdown(f"**{q['question']}**")

        current_val = st.session_state.mapping_scores.get(q["id"])
        idx = q["scale"].index(current_val) if current_val in q["scale"] else None
        chosen = st.radio(
            "Puntuación:",
            options=q["scale"],
            index=idx,
            key=f"map_score_{q['id']}",
            horizontal=True,
        )
        if chosen:
            st.session_state.mapping_scores[q["id"]] = chosen

        comment = st.text_area(
            "Amplíe su respuesta (máx. 200 palabras):",
            value=st.session_state.mapping_comments.get(q["id"], ""),
            key=f"map_comment_{q['id']}",
            placeholder="Agregue evidencias, contexto o perspectiva personal...",
            height=90,
        )
        st.session_state.mapping_comments[q["id"]] = comment
        st.markdown(word_counter_html(comment), unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)
        st.markdown("---")

    all_answered = all(
        st.session_state.mapping_scores.get(q["id"]) in q["scale"]
        for q in MAPPING_QUESTIONS
    )

    col_nav1, _, col_nav3 = st.columns([1, 2, 1])
    with col_nav1:
        if st.button("← Anterior", use_container_width=True):
            st.session_state.step = 4
            st.rerun()
    with col_nav3:
        if st.button("Siguiente →", disabled=not all_answered, use_container_width=True, type="primary"):
            st.session_state.step = 6
            st.rerun()

# ─── STEP 6: DIAGNÓSTICO FINAL ───────────────────────────────────────────────
elif st.session_state.step == 6:
    st.markdown("""
    <div class="step-header">
        <h2>🔬 Paso 7 · Diagnóstico de Cadenas Productivas y Necesidades del Ecosistema</h2>
        <p>Responda con detalle las siguientes preguntas abiertas. Su perspectiva es fundamental para el diagnóstico regional (máx. 200 palabras por respuesta).</p>
    </div>
    <div class="ref-box">
        <strong>Referencias Q1/Q2:</strong> Porter, M. E. (1998). Clusters and the new economics of competition. <em>Harvard Business Review, 76</em>(6), 77–90. · Coe, N. M. et al. (2008). Global production networks. <em>Journal of Economic Geography, 8</em>(3), 271–295. · Uyarra, E., & Flanagan, K. (2010). Understanding the innovation impacts of public procurement. <em>European Planning Studies, 18</em>(1), 123–143.
    </div>
    """, unsafe_allow_html=True)

    for q in DIAGNOSIS_QUESTIONS:
        st.markdown(f"### {q['label']}")
        st.markdown(f"*{q['question']}*")
        resp = st.text_area(
            "Su respuesta:",
            value=st.session_state.diagnosis.get(q["id"], ""),
            key=f"diag_{q['id']}",
            placeholder="Escriba aquí su respuesta (máx. 200 palabras)...",
            height=120,
        )
        st.session_state.diagnosis[q["id"]] = resp
        st.markdown(word_counter_html(resp), unsafe_allow_html=True)
        st.markdown("---")

    col_nav1, _, col_nav3 = st.columns([1, 2, 1])
    with col_nav1:
        if st.button("← Anterior", use_container_width=True):
            st.session_state.step = 5
            st.rerun()
    with col_nav3:
        if st.button("✅  Enviar respuestas", use_container_width=True, type="primary"):
            st.session_state.submitted = True
            st.rerun()

# ─── FOOTER ──────────────────────────────────────────────────────────────────
st.markdown("<br>", unsafe_allow_html=True)
st.markdown("""
<div style="border-top:1px solid #d0cbbf;padding-top:1rem;font-size:0.72rem;color:#8a8578;line-height:1.8;">
<strong style="color:#0d0f1a;font-family:'Syne',sans-serif;">IDEP — Instrumento Diagnóstico de Ecosistemas Productivos</strong><br>
Investigación Doctoral en Innovación y Productividad Regional · Modelo de la Quinta Hélice · 2025<br><br>
<strong>Bibliografía de referencia (Q1/Q2):</strong><br>
Carayannis, E. G., & Campbell, D. F. J. (2010). <em>Int. J. Social Ecology and Sustainable Development, 1</em>(1), 41–69 ·
Jacobides, M. G. et al. (2018). <em>Strategic Management Journal, 39</em>(8), 2255–2276 ·
Granstrand, O., & Holgersson, M. (2020). <em>Technovation, 90–91</em>, 102098 ·
Gereffi, G., & Fernandez-Stark, K. (2016). Global Value Chain Analysis. CGGC ·
Stam, E. (2015). <em>European Planning Studies, 23</em>(6), 1759–1762 ·
Autio, E. et al. (2018). <em>Strategic Entrepreneurship Journal, 12</em>(1), 72–95
</div>
""", unsafe_allow_html=True)
